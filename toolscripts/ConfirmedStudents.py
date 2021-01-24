# -*- coding: utf-8 -*-
# @Time    : 11/19/20 7:28 PM
# @Author  : Saptarshi
# @Email   : saptarshi.mitra@tum.de
# @File    : ConfirmedStudents.py
# @Project: eda-seminar-organization-grading

#def add(a: int, b: int) -> int:
#   return a + b

import numpy as np
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *
import sys
import argparse
import os

def create_arg_parser():
    """Creates and returns the ArgumentParser object"""

    parser = argparse.ArgumentParser(description='Gives confirmed list of students with review allocation')
    parser.add_argument('inputFile',
                    help='Path to the input xlsx file')
    parser.add_argument('-HS', dest='hauptseminar', action='store_true',
                    help='use this switch to toggle to Hauptseminar')
    parser.add_argument('-u', '--update', dest='updatepath', action='append',
                        help='use this to update an existing master file with new students from moodle')
    return parser


def read_srcfile(source_filename):
    """Import the src file from TUMonline into a Pandas dataframe"""

    src_wb = load_workbook(source_filename)
    print("The available sheets in the xlsx file")
    print(src_wb.sheetnames)
    src_sheet = src_wb.active
    print("selected sheet for data manipulation:")
    print(src_sheet)
    src_df = pd.DataFrame(src_sheet.values)     #return the sheet values as pandas Dataframe
    return src_df


def write_masterfile(write_df):
    """Create Xlsx file with current timestamp(to identify latest updated version. Then copying the pandas df in Sheets)"""
    writer = pd.ExcelWriter("OutputFiles/master_sheet_" + str(pd.datetime.now().strftime('%Y_%m_%d_%H_%M_%S')) + ".xlsx", engine='xlsxwriter')
    write_df.to_excel(writer,'Sheet1')
    writer.save()


def replace_header(input_df):
    """replace headers of the dataframe with first row of sheet"""
    new_header = input_df.iloc[0]
    input_df = input_df[1:]
    input_df.columns=new_header
    return input_df


def choose_fixed_place(modified_df):
    """This function selects the students with Status as Fixed place and filter out the column names for  those students which are required for the Master Sheet"""
    filtered_df = modified_df[(modified_df.STATUS=='Fixplatz')]
    filtered_df.reset_index(inplace=True, drop=True)
    filtered_df = filtered_df[['FAMILIENNAME','VORNAME','MATRIKELNUMMER','GESCHLECHT','E-MAIL']]
    return filtered_df


def add_columns(filtered_df):
    """This function adds additional columns for TITEL, BETREUER, VORTRAG for each student with a fixed place, the details can be manually filled after the Master Sheet is generated"""
    print("adding additional columns for TITEL, BETREUER, VORTRAG")
    extended_df = filtered_df.assign(TITEL='', BETREUER='', VORTRAG='')
    return extended_df


def add_columns_HS(extended_df):
    """Only in case of Hauptseminar, this function adds additional columns for REVIEW FÜR (AKTIVE), REVIEW VON (PASSIV)"""
    print("adding additional columns for REVIEW FÜR (AKTIVE), REVIEW VON (PASSIV)")
    extended_df = extended_df.assign(REVIEW_FÜR='', REVIEW_VON='')
    return extended_df


def shuffle_review(extended_df):
    """In case of HauptSeminar, Shuffling the review for column to assign a student report to another student"""
    print("Shuffling the review for column to assign a student report to another student")
    #uncomment this section if you want random shuffling
    # shuffled_review_column = extended_df.drop(extended_df.columns.difference(['REVIEW_FÜR']), 1)
    # shuffled_review_column['REVIEW_FÜR'] = extended_df['MATRIKELNUMMER']
    # shuffled_review_column = shuffled_review_column.sample(frac=1)
    # shuffled_review_column.reset_index(inplace=True, drop=True)
    # shuffled_df = extended_df
    # shuffled_df['REVIEW_FÜR'] = shuffled_review_column['REVIEW_FÜR']
    # shuffled_df_copy = shuffled_df

    #shift by 1 review allotement
    for index, row in extended_df.iterrows():
        if index < (extended_df.shape[0]- 1):
            row['REVIEW_FÜR'] = extended_df.iloc[index+1]['MATRIKELNUMMER']
        else:
            row['REVIEW_FÜR'] = extended_df.iloc[0]['MATRIKELNUMMER']
    for index, row in extended_df.iterrows():
        if index != 0:
            row['REVIEW_VON'] = extended_df.iloc[index-1]['MATRIKELNUMMER']
        else:
            row['REVIEW_VON'] = extended_df.iloc[extended_df.shape[0]-1]['MATRIKELNUMMER']
    return extended_df


def main():
    """Entry point for Scipt1 (Preparing the MasterSheet from TUMonline input sheet)"""
    arg_parser = create_arg_parser()
    parsed_args = arg_parser.parse_args(sys.argv[1:])   #parsing arguments after first one(self)
    if os.path.exists(parsed_args.inputFile):
        print("Input File exists")
        src_df = read_srcfile(parsed_args.inputFile)
        print(src_df.to_string())
        modified_df = replace_header(src_df)
        print(modified_df)
        filtered_df = choose_fixed_place(modified_df)
        print(filtered_df)
        extended_df = add_columns(filtered_df)
    if(parsed_args.updatepath):     #Parse the path to the manually modified file to get new entries from TUMonline
        print("Updating student information in {}".format(parsed_args.updatepath))
        des_df = read_srcfile("".join(parsed_args.updatepath))  #Read Manually modified MasterSheet
        print(des_df.to_string())
        modified_df_update = replace_header(des_df)
        print(modified_df_update)
        modified_df_update = modified_df_update.loc[:, modified_df_update.columns.notnull()]        #shaping the datframe for sheets
        modified_df_update.reset_index(inplace=True, drop=True)
        #extended_df = add_columns(filtered_df_update)
        for index, row in extended_df.iterrows():       #iterate df to check new entries
            flag_new_entry = 1
            for index_updated, row_updated in modified_df_update.iterrows():
                if extended_df.iloc[index]['MATRIKELNUMMER'] == modified_df_update.iloc[index_updated]['MATRIKELNUMMER']:
                    flag_new_entry = 0
                    break
            if flag_new_entry == 1:
                modified_df_update.loc[len(modified_df_update)] = extended_df.iloc[index]
        extended_df = modified_df_update


    # arguments = len(sys.argv) - 1
    # position = 1
    # while (arguments >= position):
    #     print("Parameter %i: %s" % (position, sys.argv[position]))
    #     position = position + 1
    # source_filename = sys.argv[1]

    #if(arguments==2 and sys.argv[2]=='-HS'):
    if (parsed_args.hauptseminar):      #check if -HS key is used, then extra features added to MasterSheet for Hauptseminar
        extended_df = add_columns_HS(extended_df)
        extended_df = shuffle_review(extended_df)
    write_masterfile(extended_df)

if __name__ == "__main__":
    main()
