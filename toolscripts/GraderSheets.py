# -*- coding: utf-8 -*-
# @Time    : 12/10/20 5:30 PM
# @Author  : Saptarshi
# @Email   : saptarshi.mitra@tum.de
# @File    : GraderSheets.py
# @Project: eda-seminar-organization-grading

import numpy as np
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *
import sys
import argparse
import os

def create_arg_parser():
    """Creates and returns the ArgumentParser object"""

    parser = argparse.ArgumentParser(description='Gives confirmed list of students with review allocation')
    parser.add_argument('inputFile',
                    help='Path to the input xlsx file')
    parser.add_argument('-HS', dest='hauptseminar', action='store_true',
                    help='use this switch to toggle to Hauptseminar')
    parser.add_argument('-u', '--update', dest='updatepath', action='append',
                        help='use this to update an existing master file with new students from moodle')
    return parser

def read_srcfile(source_filename):
    """Import the src file from TUMonline into a Pandas dataframe"""
    src_wb = load_workbook(source_filename)
    print("The available sheets in the xlsx file")
    print(src_wb.sheetnames)
    src_sheet = src_wb.active
    print("selected sheet for data manipulation:")
    print(src_sheet)
    src_df = pd.DataFrame(src_sheet.values)
    return src_df




def replace_header(input_df):
    """replace headers of the dataframe with first row of sheet"""
    new_header = input_df.iloc[0]
    input_df = input_df[1:]
    input_df.columns=new_header
    return input_df




def generate_supervisor_files(modified_df):
    """Create an individual file with supervisor name, then fill it which required pages from template with customized cells, filled by data from MasterSheets"""
    supervisor_list = modified_df.BETREUER.unique()
    print("Here are the list of supervisors")
    print(supervisor_list)
    for supervisor in supervisor_list:      #iterate through all unique supervisors and add pages
        print(supervisor)
        writer = pd.ExcelWriter("OutputFiles/GraderSheets/" + supervisor + "_GradingSheetSeminar" + ".xlsx",engine='xlsxwriter')
        overview_dataframe = add_overview_sheet(supervisor)
        overview_dataframe.to_excel(writer, 'Overview')
        paper_grading_dataframe = add_paper_grading_sheet(supervisor, modified_df)
        paper_grading_dataframe.to_excel(writer, 'Paper Grading')
        review_grading_dataframe = add_review_grading_sheet(supervisor, modified_df)
        review_grading_dataframe.to_excel(writer, 'Review Grading')
        presentations_dataframe = add_presentations_sheet(supervisor, modified_df)
        presentations_dataframe.to_excel(writer, 'Presentations')
        writer.save()


def add_overview_sheet(supervisor):
    """Adding the overview sheet with generic instructions and name of the supervisor"""
    src_grading_wb = load_workbook('DataSources/Foik_GradingSheetSeminar.xlsx')
    print("The available sheets in the xlsx file")
    print(src_grading_wb.sheetnames)
    src_sheet = src_grading_wb.active
    print("selected sheet for data manipulation:")
    print(src_sheet)
    src_sheet['C2'] = supervisor
    overview_df = pd.DataFrame(src_sheet.values)
    return overview_df


def add_paper_grading_sheet(supervisor, modified_df):
    """Adding the Paper Grading sheet with all supervised student Papers and grading metrics"""
    src_grading_wb = load_workbook('DataSources/Foik_GradingSheetSeminar.xlsx')
    print("The available sheets in the xlsx file")
    print(src_grading_wb.sheetnames)
    src_sheet = src_grading_wb["Paper Grading"]
    print("selected sheet for data manipulation:")
    print(src_sheet)
    #src_sheet['E9'] = supervisor
    paper_grading_df = pd.DataFrame(src_sheet.values)
    paper_grading_df = paper_grading_df[:6]     # taking only first rows from template
    supervision_df = modified_df[(modified_df.BETREUER == supervisor)]
    supervision_df = supervision_df[['FAMILIENNAME', 'VORNAME', 'MATRIKELNUMMER', 'TITEL']]
    for index, row in supervision_df.iterrows():
        paper_grading_df.loc[len(paper_grading_df), 0] = "Paper "+ str(index+1)
        paper_grading_df.loc[len(paper_grading_df)-1,1] = row.TITEL
        paper_grading_df.loc[len(paper_grading_df), 0] = row.FAMILIENNAME
        paper_grading_df.loc[len(paper_grading_df)-1, 1] = row.VORNAME
        paper_grading_df.loc[len(paper_grading_df)-1, 2] = row.MATRIKELNUMMER
        paper_grading_df.loc[len(paper_grading_df) - 1, 3] = "Advisor:"
        paper_grading_df.loc[len(paper_grading_df) - 1, 4] = supervisor
        paper_grading_df.loc[len(paper_grading_df), 0] = "Points:"
        paper_grading_df.loc[len(paper_grading_df)-1, 1] = "na"
        paper_grading_df.loc[len(paper_grading_df), 0] = "Comments:"
        paper_grading_df.loc[len(paper_grading_df) - 1, 1] = "na"
        paper_grading_df.loc[len(paper_grading_df)] = ""
        paper_grading_df.loc[len(paper_grading_df)] = ""
        print(row)

    return paper_grading_df


def add_review_grading_sheet(supervisor, modified_df):
    """Adding the Review Grading sheet with all students who reviewed the Papers authored by the students who were supervised by the him/her and grading metrics"""
    src_grading_wb = load_workbook('DataSources/Foik_GradingSheetSeminar.xlsx')
    print("The available sheets in the xlsx file")
    print(src_grading_wb.sheetnames)
    src_sheet = src_grading_wb["Review Grading"]
    print("selected sheet for data manipulation:")
    print(src_sheet)
    #src_sheet['E9'] = supervisor
    review_grading_df = pd.DataFrame(src_sheet.values)
    reviewer_interaction_df = review_grading_df[8:18]
    review_grading_df = review_grading_df[:5]     # taking only first rows from template
    supervision_df = modified_df[(modified_df.BETREUER == supervisor)]
    supervision_df = supervision_df[['FAMILIENNAME', 'VORNAME', 'MATRIKELNUMMER', 'TITEL', 'REVIEW_VON']]
    for index, row in supervision_df.iterrows():
        review_grading_df.loc[len(review_grading_df), 0] = "Review for Paper "+ str(index+1)
        review_grading_df.loc[len(review_grading_df)-1,2] = "Title: "+ row.TITEL
        review_grading_df.loc[len(review_grading_df),0] = "Author:"
        review_grading_df.loc[len(review_grading_df)-1, 1] = row.FAMILIENNAME
        review_grading_df.loc[len(review_grading_df)-1, 2] = row.VORNAME
        review_grading_df.loc[len(review_grading_df)-1, 3] = row.MATRIKELNUMMER
        review_grading_df.loc[len(review_grading_df)-1, 4] = "Advisor:"
        review_grading_df.loc[len(review_grading_df) - 1, 5] = supervisor
        for indexreview, rowreview in modified_df.iterrows():
            if rowreview.REVIEW_FÜR == row.MATRIKELNUMMER:
                review_grading_df.loc[len(review_grading_df), 0] = "Reviewer:"
                review_grading_df.loc[len(review_grading_df)-1, 1] = rowreview.FAMILIENNAME
                review_grading_df.loc[len(review_grading_df) - 1, 2] = rowreview.VORNAME
                review_grading_df.loc[len(review_grading_df) - 1, 3] = rowreview.MATRIKELNUMMER
                review_grading_df = pd.concat([review_grading_df,reviewer_interaction_df],axis=0,ignore_index=True)
        review_grading_df.loc[len(review_grading_df)] = ""
        review_grading_df.loc[len(review_grading_df)] = ""
    return review_grading_df


def add_presentations_sheet(supervisor, modified_df):
    """Adding the Presentations sheet with all Presentations according to the talk order on the presentation day"""
    src_grading_wb = load_workbook('DataSources/Foik_GradingSheetSeminar.xlsx')
    print("The available sheets in the xlsx file")
    print(src_grading_wb.sheetnames)
    src_sheet = src_grading_wb["Presentations"]
    print("selected sheet for data manipulation:")
    print(src_sheet)
    review_grading_df = pd.DataFrame(src_sheet.values)
    reviewer_interaction_df = review_grading_df[3:13]
    review_grading_df = review_grading_df[:1]     # taking only first rows from template

    for talk_number in range(len(modified_df)):
        for indextalk, rowtalk in modified_df.iterrows():
            if rowtalk.VORTRAG == (talk_number+1):
                review_grading_df.loc[len(review_grading_df), 0] = "Talk " + str(talk_number + 1)
                review_grading_df.loc[len(review_grading_df) - 1, 1] = rowtalk.TITEL
                review_grading_df.loc[len(review_grading_df), 0] = "Speaker:"
                review_grading_df.loc[len(review_grading_df) - 1, 1] = rowtalk.VORNAME
                review_grading_df.loc[len(review_grading_df) - 1, 2] = rowtalk.FAMILIENNAME
                review_grading_df.loc[len(review_grading_df) - 1, 3] = rowtalk.MATRIKELNUMMER
                review_grading_df.loc[len(review_grading_df) - 1, 4] = "Advisor:"
                review_grading_df.loc[len(review_grading_df) - 1, 5] = rowtalk.BETREUER
                review_grading_df = pd.concat([review_grading_df, reviewer_interaction_df], axis=0, ignore_index=True)

        review_grading_df.loc[len(review_grading_df)] = ""
        review_grading_df.loc[len(review_grading_df)] = ""

    return review_grading_df








def main():
    """Entry point for Scipt2 (Preparing the GradingSheets for each supervisors from MasterSheet produced by Script1"""
    arg_parser = create_arg_parser()
    parsed_args = arg_parser.parse_args(sys.argv[1:])        #parsing arguments after first one(self)
    if os.path.exists(parsed_args.inputFile):
        print("Input File exists")
        #src_df = read_srcfile(parsed_args.inputFile)
        src_df = read_srcfile("".join(parsed_args.inputFile))
        print(src_df.to_string())
        modified_df = replace_header(src_df)
        print(modified_df)
        modified_df = modified_df.loc[:, modified_df.columns.notnull()]     #shaping the dataframe for sheets
        modified_df.reset_index(inplace=True, drop=True)

    generate_supervisor_files(modified_df)

    # if(parsed_args.updatepath):
    #     print("Updating student information in {}".format(parsed_args.updatepath))
    #     des_df = read_srcfile("".join(parsed_args.updatepath))
    #     print(des_df.to_string())
    #     modified_df_update = replace_header(des_df)
    #     print(modified_df_update)
    #     modified_df_update = modified_df_update.loc[:, modified_df_update.columns.notnull()]
    #     modified_df_update.reset_index(inplace=True, drop=True)
    #     #extended_df = add_columns(filtered_df_update)
    #     for index, row in extended_df.iterrows():
    #         flag_new_entry = 1
    #         for index_updated, row_updated in modified_df_update.iterrows():
    #             if extended_df.iloc[index]['MATRIKELNUMMER'] == modified_df_update.iloc[index_updated]['MATRIKELNUMMER']:
    #                 flag_new_entry = 0
    #                 break
    #         if flag_new_entry == 1:
    #             modified_df_update.loc[len(modified_df_update)] = extended_df.iloc[index]
    #     extended_df = modified_df_update




    # if (parsed_args.hauptseminar):
    #     extended_df = add_columns_HS(extended_df)
    #     extended_df = shuffle_review(extended_df)


    #write_masterfile(extended_df)

if __name__ == "__main__":
    main()
